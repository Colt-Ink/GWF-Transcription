# api_communication.py

import requests
import time
import openpyxl
import pprint
from api_secrets import API_KEY_ASSEMBLYAI
from tqdm import tqdm
from tqdm import trange

# upload
upload_endpoint = "https://api.assemblyai.com/v2/upload"
transcript_endpoint = "https://api.assemblyai.com/v2/transcript"

headers_auth_only = {'authorization': API_KEY_ASSEMBLYAI}

headers = {
    "authorization": API_KEY_ASSEMBLYAI,
    "content-type": "application/json"
}

CHUNK_SIZE = 5_242_880  # 5MB

def upload(url, title):
    """
    Uploads audio file to AssemblyAI
    """
    data = {
        "audio_src": url,
        "metadata": {
            "title": title
        }
    }
    response = requests.post(upload_endpoint, json=data, headers=headers)
    return response.json()['upload_url']

def request_transcript(upload_url):
    """
    Requests transcript from AssemblyAI
    """
    data = {
        "audio_url": upload_url,
        "model": "default",
        "speaker_channels": "combined",
        "speaker_labels": True,
        "confidence_threshold": 0.5,
        "auto_highlights": True,
        "iab_categories": True,
        "summarization": True,
        "auto_chapters": True,
        "entity_detection": True
    }
    response = requests.post(transcript_endpoint, json=data, headers=headers)
    return response.json()['id']

def poll_transcript(transcript_id):
    """
    Polls transcript endpoint for status
    """
    polling_endpoint = transcript_endpoint + '/' + transcript_id
    response = requests.get(polling_endpoint, headers=headers_auth_only)
    return response.json()

def request_paragraphs(transcript_id):
    """
    Requests paragraphs from AssemblyAI
    """
    paragraphs_endpoint = transcript_endpoint + '/' + transcript_id + '/paragraphs'
    response = requests.get(paragraphs_endpoint, headers=headers_auth_only)
    return response.json()

def request_srt(transcript_id):
    """
    Requests srt from AssemblyAI
    """
    srt_endpoint = transcript_endpoint + '/' + transcript_id + '/srt'
    response = requests.get(srt_endpoint, headers=headers_auth_only)
    return response.json()

def save_transcript(url, title):
    """
    Saves transcript to Excel file with sub-sheets for Auto-Highlights, IAB Categories, Summarization, Chapters, and Entities
    """
    upload_url = upload(url, title)
    transcript_id = request_transcript(upload_url)
    response = poll_transcript(transcript_id)
    while response['status'] != 'completed':
        response = poll_transcript(transcript_id)
        time.sleep(1)
    paragraphs = request_paragraphs(transcript_id)
    srt = request_srt(transcript_id)
    wb = openpyxl.Workbook()
    wb.create_sheet('Transcript')
    wb.create_sheet('Auto-Highlights')
    wb.create_sheet('IAB Categories')
    wb.create_sheet('Summarization')
    wb.create_sheet('Chapters')
    wb.create_sheet('Entities')
    wb.save(title + '.xlsx')
    wb = openpyxl.load_workbook(title + '.xlsx')
    ws = wb['Transcript']
    ws.cell(row=1, column=1).value = 'Speaker'
    ws.cell(row=1, column=2).value = 'Text'
    ws.cell(row=1, column=3).value = 'Start Time'
    ws.cell(row=1, column=4).value = 'End Time'
    ws.cell(row=1, column=5).value = 'Confidence'
    ws.cell(row=1, column=6).value = 'Highlight'
    ws.cell(row=1, column=7).value = 'IAB Category'
    ws.cell(row=1, column=8).value = 'Summarization'
    ws.cell(row=1, column=9).value = 'Chapter'
    ws.cell(row=1, column=10).value = 'Entity'
    ws.cell(row=1, column=11).value = 'Entity Type'
    ws.cell(row=1, column=12).value = 'Entity Subtype'
    for i in range(len(paragraphs)):
        ws.cell(row=i+2, column=1).value = paragraphs[i]['speaker']
        ws.cell(row=i+2, column=2).value = paragraphs[i]['text']
        ws.cell(row=i+2, column=3).value = paragraphs[i]['start_time']
        ws.cell(row=i+2, column=4).value = paragraphs[i]['end_time']
        ws.cell(row=i+2, column=5).value = paragraphs[i]['confidence']
        ws.cell(row=i+2, column=6).value = paragraphs[i]['highlight']
        ws.cell(row=i+2, column=7).value = paragraphs[i]['iab_category']
        ws.cell(row=i+2, column=8).value = paragraphs[i]['summarization']
        ws.cell(row=i+2, column=9).value = paragraphs[i]['chapter']
        ws.cell(row=i+2, column=10).value = paragraphs[i]['entity']
        ws.cell(row=i+2, column=11).value = paragraphs[i]['entity_type']
        ws.cell(row=i+2, column=12).value = paragraphs[i]['entity_subtype']
    wb.save(title + '.xlsx')
    wb.close()
    wb = openpyxl.load_workbook(title + '.xlsx')
    ws = wb['Auto-Highlights']
    ws.cell(row=1, column=1).value = 'Speaker'
    ws.cell(row=1, column=2).value = 'Text'
    ws.cell(row=1, column=3).value = 'Start Time'
    ws.cell(row=1, column=4).value = 'End Time'
    ws.cell(row=1, column=5).value = 'Confidence'
    ws.cell(row=1, column=6).value = 'Highlight'
    ws.cell(row=1, column=7).value = 'IAB Category'
    ws.cell(row=1, column=8).value = 'Summarization'
    ws.cell(row=1, column=9).value = 'Chapter'
    ws.cell(row=1, column=10).value = 'Entity'
    ws.cell(row=1, column=11).value = 'Entity Type'
    ws.cell(row=1, column=12).value = 'Entity Subtype'
    for i in range(len(paragraphs)):
        if paragraphs[i]['highlight'] == True:
            ws.cell(row=i+2, column=1).value = paragraphs[i]['speaker']
            ws.cell(row=i+2, column=2).value = paragraphs[i]['text']
            ws.cell(row=i+2, column=3).value = paragraphs[i]['start_time']
            ws.cell(row=i+2, column=4).value = paragraphs[i]['end_time']
            ws.cell(row=i+2, column=5).value = paragraphs[i]['confidence']
            ws.cell(row=i+2, column=6).value = paragraphs[i]['highlight']
            ws.cell(row=i+2, column=7).value = paragraphs[i]['iab_category']
            ws.cell(row=i+2, column=8).value = paragraphs[i]['summarization']
            ws.cell(row=i+2, column=9).value = paragraphs[i]['chapter']
            ws.cell(row=i+2, column=10).value = paragraphs[i]['entity']
            ws.cell(row=i+2, column=11).value = paragraphs[i]['entity_type']
            ws.cell(row=i+2, column=12).value = paragraphs[i]['entity_subtype']
    wb.save(title + '.xlsx')
    wb.close()
    wb = openpyxl.load_workbook(title + '.xlsx')
    ws = wb['IAB Categories']
    ws.cell(row=1, column=1).value = 'Speaker'
    ws.cell(row=1, column=2).value = 'Text'
    ws.cell(row=1, column=3).value = 'Start Time'
    ws.cell(row=1, column=4).value = 'End Time'
    ws.cell(row=1, column=5).value = 'Confidence'
    ws.cell(row=1, column=6).value = 'Highlight'
    ws.cell(row=1, column=7).value = 'IAB Category'
    ws.cell(row=1, column=8).value = 'Summarization'
    ws.cell(row=1, column=9).value = 'Chapter'
    ws.cell(row=1, column=10).value = 'Entity'
    ws.cell(row=1, column=11).value = 'Entity Type'
    ws.cell(row=1, column=12).value = 'Entity Subtype'
    for i in range(len(paragraphs)):
        if paragraphs[i]['iab_category'] != None:
            ws.cell(row=i+2, column=1).value = paragraphs[i]['speaker']
            ws.cell(row=i+2, column=2).value = paragraphs[i]['text']
            ws.cell(row=i+2, column=3).value = paragraphs[i]['start_time']
            ws.cell(row=i+2, column=4).value = paragraphs[i]['end_time']
            ws.cell(row=i+2, column=5).value = paragraphs[i]['confidence']
            ws.cell(row=i+2, column=6).value = paragraphs[i]['highlight']
            ws.cell(row=i+2, column=7).value = paragraphs[i]['iab_category']
            ws.cell(row=i+2, column=8).value = paragraphs[i]['summarization']
            ws.cell(row=i+2, column=9).value = paragraphs[i]['chapter']
            ws.cell(row=i+2, column=10).value = paragraphs[i]['entity']
            ws.cell(row=i+2, column=11).value = paragraphs[i]['entity_type']
            ws.cell(row=i+2, column=12).value = paragraphs[i]['entity_subtype']
    wb.save(title + '.xlsx')
    wb.close()
    wb = openpyxl.load_workbook(title + '.xlsx')
    ws = wb['Summarization']
    ws.cell(row=1, column=1).value = 'Speaker'
    ws.cell(row=1, column=2).value = 'Text'
    ws.cell(row=1, column=3).value = 'Start Time'
    ws.cell(row=1, column=4).value = 'End Time'
    ws.cell(row=1, column=5).value = 'Confidence'
    ws.cell(row=1, column=6).value = 'Highlight'
    ws.cell(row=1, column=7).value = 'IAB Category'
    ws.cell(row=1, column=8).value = 'Summarization'
    ws.cell(row=1, column=9).value = 'Chapter'
    ws.cell(row=1, column=10).value = 'Entity'
    ws.cell(row=1, column=11).value = 'Entity Type'
    ws.cell(row=1, column=12).value = 'Entity Subtype'
    for i in range(len(paragraphs)):
        if paragraphs[i]['summarization'] == True:
            ws.cell(row=i+2, column=1).value = paragraphs[i]['speaker']
            ws.cell(row=i+2, column=2).value = paragraphs[i]['text']
            ws.cell(row=i+2, column=3).value = paragraphs[i]['start_time']
            ws.cell(row=i+2, column=4).value = paragraphs[i]['end_time']
            ws.cell(row=i+2, column=5).value = paragraphs[i]['confidence']
            ws.cell(row=i+2, column=6).value = paragraphs[i]['highlight']
            ws.cell(row=i+2, column=7).value = paragraphs[i]['iab_category']
            ws.cell(row=i+2, column=8).value = paragraphs[i]['summarization']
            ws.cell(row=i+2, column=9).value = paragraphs[i]['chapter']
            ws.cell(row=i+2, column=10).value = paragraphs[i]['entity']
            ws.cell(row=i+2, column=11).value = paragraphs[i]['entity_type']
            ws.cell(row=i+2, column=12).value = paragraphs[i]['entity_subtype']
    wb.save(title + '.xlsx')
    wb.close()
    wb = openpyxl.load_workbook(title + '.xlsx')
    ws = wb['Chapters']
    ws.cell(row=1, column=1).value = 'Speaker'
    ws.cell(row=1, column=2).value = 'Text'
    ws.cell(row=1, column=3).value = 'Start Time'
    ws.cell(row=1, column=4).value = 'End Time'
    ws.cell(row=1, column=5).value = 'Confidence'
    ws.cell(row=1, column=6).value = 'Highlight'
    ws.cell(row=1, column=7).value = 'IAB Category'
    ws.cell(row=1, column=8).value = 'Summarization'
    ws.cell(row=1, column=9).value = 'Chapter'
    ws.cell(row=1, column=10).value = 'Entity'
    ws.cell(row=1, column=11).value = 'Entity Type'
    ws.cell(row=1, column=12).value = 'Entity Subtype'
    for i in range(len(paragraphs)):
        if paragraphs[i]['chapter'] != None:
            ws.cell(row=i+2, column=1).value = paragraphs[i]['speaker']
            ws.cell(row=i+2, column=2).value = paragraphs[i]['text']
            ws.cell(row=i+2, column=3).value = paragraphs[i]['start_time']
            ws.cell(row=i+2, column=4).value = paragraphs[i]['end_time']
            ws.cell(row=i+2, column=5).value = paragraphs[i]['confidence']
            ws.cell(row=i+2, column=6).value = paragraphs[i]['highlight']
            ws.cell(row=i+2, column=7).value = paragraphs[i]['iab_category']
            ws.cell(row=i+2, column=8).value = paragraphs[i]['summarization']
            ws.cell(row=i+2, column=9).value = paragraphs[i]['chapter']
            ws.cell(row=i+2, column=10).value = paragraphs[i]['entity']
            ws.cell(row=i+2, column=11).value = paragraphs[i]['entity_type']
            ws.cell(row=i+2, column=12).value = paragraphs[i]['entity_subtype']
    wb.save(title + '.xlsx')
    wb.close()
    wb = openpyxl.load_workbook(title + '.xlsx')
    ws = wb['Entities']
    ws.cell(row=1, column=1).value = 'Speaker'
    ws.cell(row=1, column=2).value = 'Text'
    ws.cell(row=1, column=3).value = 'Start Time'
    ws.cell(row=1, column=4).value = 'End Time'
    ws.cell(row=1, column=5).value = 'Confidence'
    ws.cell(row=1, column=6).value = 'Highlight'
    ws.cell(row=1, column=7).value = 'IAB Category'
    ws.cell(row=1, column=8).value = 'Summarization'
    ws.cell(row=1, column=9).value = 'Chapter'
    ws.cell(row=1, column=10).value = 'Entity'
    ws.cell(row=1, column=11).value = 'Entity Type'
    ws.cell(row=1, column=12).value = 'Entity Subtype'
    for i in range(len(paragraphs)):
        if paragraphs[i]['entity'] != None:
            ws.cell(row=i+2, column=1).value = paragraphs[i]['speaker']
            ws.cell(row=i+2, column=2).value = paragraphs[i]['text']
            ws.cell(row=i+2, column=3).value = paragraphs[i]['start_time']
            ws.cell(row=i+2, column=4).value = paragraphs[i]['end_time']
            ws.cell(row=i+2, column=5).value = paragraphs[i]['confidence']
            ws.cell(row=i+2, column=6).value = paragraphs[i]['highlight']
            ws.cell(row=i+2, column=7).value = paragraphs[i]['iab_category']
            ws.cell(row=i+2, column=8).value = paragraphs[i]['summarization']
            ws.cell(row=i+2, column=9).value = paragraphs[i]['chapter']
            ws.cell(row=i+2, column=10).value = paragraphs[i]['entity']
            ws.cell(row=i+2, column=11).value = paragraphs[i]['entity_type']
            ws.cell(row=i+2, column=12).value = paragraphs[i]['entity_subtype']
    wb.save(title + '.xlsx')
    wb.close()
    print("Transcript saved to " + title + ".xlsx")