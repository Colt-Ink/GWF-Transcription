import requests
import time
import openpyxl
from api_secrets import API_KEY_ASSEMBLYAI

# upload
upload_endpoint = "https://api.assemblyai.com/v2/upload"
transcript_endpoint = "https://api.assemblyai.com/v2/transcript"

headers_auth_only = {'authorization': API_KEY_ASSEMBLYAI}

headers = {
    "authorization": API_KEY_ASSEMBLYAI,
    "content-type": "application/json"
}

CHUNK_SIZE = 5_242_880  # 5MB

def upload(filename):
    def read_file(filename):
        with open(filename, 'rb') as f:
            while True:
                data = f.read(CHUNK_SIZE)
                if not data:
                    break
                yield data

    upload_response = requests.post(upload_endpoint, headers=headers_auth_only, data=read_file(filename))
    return upload_response.json()['upload_url']

# transcribe

def transcribe(audio_url):
    transcript_request = {
        'audio_url': audio_url,
        'auto_highlights': True, # Highlights
        'iab_categories': True, # Topics
        'summarization': True, # Summarization
        'auto_chapters': True, # Chapters
        'entity_detection': True # Entities
    }

    transcript_response = requests.post(transcript_endpoint, json=transcript_request, headers=headers)
    return transcript_response.json()

        
def poll(transcript_id):
    polling_endpoint = transcript_endpoint + '/' + transcript_id['id']
    polling_response = requests.get(polling_endpoint, headers=headers)
    return polling_response.json()


def get_transcription_result_url(url):
    transcribe_id = transcribe(url)
    while True:
        data = poll(transcribe_id)
        if data['status'] == 'completed':
            return data, None
        elif data['status'] == 'error':
            return data, data['error']
            
        print("waiting for 30 seconds")
        time.sleep(30)
        
def request_paragraphs(data):
    paragraphs_endpoint = transcript_endpoint + '/' + data['id'] + '/paragraphs'
    paragraphs_response = requests.get(paragraphs_endpoint, headers=headers)
    return paragraphs_response.json()

def request_srt(data):
    srt_endpoint = transcript_endpoint + '/' + data['id'] + '/srt'
    srt_response = requests.get(srt_endpoint, headers=headers)
    return srt_response.json()

def request_vtt(data):
    vtt_endpoint = transcript_endpoint + '/' + data['id'] + '/vtt'
    vtt_response = requests.get(vtt_endpoint, headers=headers)
    return vtt_response.json()

def save_transcript(url, title):
    data, error = get_transcription_result_url(url)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Transcript"
    ws.append(data)
    
    ws = wb.create_sheet(title="Paragraphs")
    ws.append(request_paragraphs(data))
    
    ws = wb.create_sheet(title="VTT")
    ws.append(request_vtt(data))
    
    ws = wb.create_sheet(title="SRT")
    ws.append(request_srt(data))
    
    ws = wb.create_sheet(title="Highlights")
    ws.append(data['highlights'])
    
    ws = wb.create_sheet(title="Topics")
    ws.append(data['topics'])
    
    ws = wb.create_sheet(title="Summarization")
    ws.append(data['summarization'])
    
    ws = wb.create_sheet(title="Chapters")
    ws.append(data['chapters'])
    
    ws = wb.create_sheet(title="Entities")
    ws.append(data['entities'])
    
    wb.save(title + '.xlsx')
    
    if data:
        print('Transcript saved')
    elif error:
        print("Error!!!", error)